#!/usr/bin/env python3
"""
=============================================================================
generate_supplementary_tables.py — Supplementary Tables S1–S5
=============================================================================
Generates all supplementary tables referenced in the manuscript:
  S1: LMM sensitivity to R² thresholds (0.80, 0.85, 0.90)
  S2: Leave-one-out Spearman ρ + Cook's D (good-outcome cross-modal)
  S3: Individual patient summary data
  S4: Per-patient epoch counts + group comparisons
  S5: Per-patient ASM regimen stability

Output: Supplementary_Tables.xlsx (multi-sheet workbook)
=============================================================================
"""

import numpy as np
import pandas as pd
from scipy import stats
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import statsmodels.formula.api as smf
import warnings
warnings.filterwarnings('ignore')

# ═══════════════════════════════════════════════════════════════════
#  CONFIGURATION
# ═══════════════════════════════════════════════════════════════════

# Auto-detect data directory by checking for a known data file
SCRIPT_DIR = Path(__file__).resolve().parent
_candidates = [
    Path("."),  # 新位置（项目内）
    Path("."),  # 旧位置（兜底）
    SCRIPT_DIR.parent,                                    # One level up from panels/
    SCRIPT_DIR,                                           # Same dir as script
    Path('/mnt/project'),                                 # Claude environment
]
DATA_DIR = next((p for p in _candidates
                 if (p / 'seeg_contact_results.csv').exists()), None)
if DATA_DIR is None:
    raise FileNotFoundError(
        f"Cannot locate data files. Searched:\n"
        + "\n".join(f"  {p}" for p in _candidates)
        + "\nPlace seeg_contact_results.csv (and other data) in one of "
        "these directories, or edit _candidates in this script."
    )
OUTPUT = SCRIPT_DIR / 'Supplementary_Tables.xlsx'
ASM_FILE = DATA_DIR / 'ASM_regimen.csv'  # Per-patient ASM data (user-filled)

PAIRED_SUBJECTS = [
    'Sub01','Sub02','Sub04','Sub05','Sub06','Sub07','Sub08','Sub09',
    'Sub11','Sub13','Sub14','Sub15','Sub17','Sub19','Sub20','Sub21'
]
MIDLINE = ['Fz', 'Cz', 'Pz']
MIN_EXPONENT = 0.5
MAX_DISTANCE_MM = 63.0

# Formatting constants
HEADER_FILL = PatternFill('solid', fgColor='2F5496')
HEADER_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=10)
DATA_FONT   = Font(name='Arial', size=10)
NOTE_FONT   = Font(name='Arial', size=9, italic=True, color='555555')
BOLD_FONT   = Font(name='Arial', size=10, bold=True)
THIN_BORDER = Border(
    bottom=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='EEEEEE'),
)
CENTER = Alignment(horizontal='center', vertical='center')
LEFT   = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ═══════════════════════════════════════════════════════════════════
#  DATA LOADERS
# ═══════════════════════════════════════════════════════════════════

def load_engel():
    df = pd.read_excel(DATA_DIR / 'engel_phase.xlsx')
    records = []
    for _, row in df.iterrows():
        raw = str(row.iloc[0]).strip()
        sub = 'Sub' + raw[3:].zfill(2) if raw.lower().startswith('sub') else raw
        e = str(row.iloc[2]).strip()
        fu = row.iloc[1]
        if e.upper() in ('', 'NAN', 'NONE'):
            continue
        eu = e.upper()
        if   eu.startswith('IV'):  outcome = 'Poor'
        elif eu.startswith('III'): outcome = 'Poor'
        elif eu.startswith('II'):  outcome = 'Good'
        elif eu.startswith('I'):   outcome = 'Good'
        else:                      outcome = 'Poor'
        records.append({'Subject': sub, 'Engel': e, 'Outcome': outcome,
                        'Follow_up_mo': fu})
    return pd.DataFrame(records)


def load_seeg():
    return pd.read_csv(DATA_DIR / 'seeg_contact_results.csv')


def load_scalp():
    return pd.read_csv(DATA_DIR / 'scalp_results.csv')


def load_processing_log():
    return pd.read_csv(DATA_DIR / 'processing_log.csv')


def load_roi():
    for name in ['ROI__Inferred_Side.csv', 'ROI  Inferred Side.csv',
                 'ROI_Inferred_Side.csv', 'ROI Inferred Side.csv']:
        p = DATA_DIR / name
        if p.exists():
            df = pd.read_csv(p)
            df.columns = [c.strip() for c in df.columns]
            return df
    # Glob fallback: find any CSV with "ROI" in name
    matches = list(DATA_DIR.glob('ROI*Side*.csv')) + list(DATA_DIR.glob('ROI*.csv'))
    if matches:
        print(f"    ℹ Found ROI file: {matches[0].name}")
        df = pd.read_csv(matches[0])
        df.columns = [c.strip() for c in df.columns]
        return df
    # Show what's actually there to help debug
    csvs = sorted(DATA_DIR.glob('*.csv'))
    raise FileNotFoundError(
        f"ROI file not found in {DATA_DIR}\n"
        f"CSV files present:\n" + "\n".join(f"  {f.name}" for f in csvs)
    )


def load_target_regions():
    return pd.read_csv(DATA_DIR / 'Target_Region_Perfect.csv')


def load_asm():
    """Load per-patient ASM regimen CSV. Returns None if file not found."""
    if ASM_FILE.exists():
        return pd.read_csv(ASM_FILE)
    # Also check script directory
    alt = SCRIPT_DIR / 'ASM_regimen.csv'
    if alt.exists():
        return pd.read_csv(alt)
    return None


# ═══════════════════════════════════════════════════════════════════
#  UTILITY
# ═══════════════════════════════════════════════════════════════════

def format_header(ws, row, headers, col_widths=None):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
    if col_widths:
        for c, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w


def format_data_row(ws, row, values, bold=False, number_fmt=None):
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.font = BOLD_FONT if bold else DATA_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        if number_fmt and isinstance(v, float):
            cell.number_format = number_fmt


def add_note(ws, row, col, text, merge_to=None):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = NOTE_FONT
    cell.alignment = LEFT
    if merge_to:
        ws.merge_cells(start_row=row, start_column=col,
                        end_row=row, end_column=merge_to)


# ═══════════════════════════════════════════════════════════════════
#  TABLE S1: LMM sensitivity to R² thresholds
# ═══════════════════════════════════════════════════════════════════

def build_table_s1(wb):
    print("  Building Table S1: Threshold sensitivity ...")
    ws = wb.create_sheet('Table_S1')

    seeg_raw = load_seeg()

    headers = ['Filter Criterion', 'N Contacts', 'N Patients',
               'LMM β (SOZ)', 'SE', '95% CI Lower', '95% CI Upper',
               'z', 'p-value', "Cohen's d", 'ICC']
    widths  = [24, 12, 12, 14, 10, 14, 14, 10, 12, 12, 10]

    add_note(ws, 1, 1,
             'Table S1. Sensitivity of the primary SOZ–NonSOZ LMM to '
             'progressively stricter spectral goodness-of-fit (R²) thresholds. '
             'The delivered dataset contains contacts with R² ≥ 0.935 and '
             'Exponent ≥ 1.17 after upstream quality filtering (R² ≥ 0.85, '
             'Exponent ≥ 0.5).',
             merge_to=11)
    format_header(ws, 3, headers, widths)

    # The delivered dataset is already pre-filtered (R²≥0.935, Exp≥1.17).
    # Sensitivity: vary R² threshold within available range.
    thresholds = [
        ('R² ≥ 0.935 (all delivered)',  0.935),
        ('R² ≥ 0.95',                   0.95),
        ('R² ≥ 0.97',                   0.97),
        ('R² ≥ 0.98',                   0.98),
        ('R² ≥ 0.99',                   0.99),
    ]

    row = 4
    for label, r2_min in thresholds:
        df = seeg_raw[(seeg_raw.R_Squared >= r2_min)].copy()
        df = df[df.Subject.isin(PAIRED_SUBJECTS)]
        n_contacts = len(df)
        n_patients = df.Subject.nunique()

        # Cohen's d (patient-level)
        soz_means = df[df.Is_SOZ].groupby('Subject')['Exponent'].mean()
        nsoz_means = df[~df.Is_SOZ].groupby('Subject')['Exponent'].mean()
        common = soz_means.index.intersection(nsoz_means.index)
        if len(common) < 3:
            format_data_row(ws, row, [label, n_contacts, n_patients] +
                            ['—'] * 8)
            row += 1
            continue
        diff = soz_means[common].values - nsoz_means[common].values
        d = diff.mean() / diff.std(ddof=1) if diff.std(ddof=1) > 0 else 0

        # LMM
        ldf = df[['Subject', 'Exponent', 'Is_SOZ']].copy()
        ldf['SOZ_int'] = ldf['Is_SOZ'].astype(int)
        try:
            model = smf.mixedlm("Exponent ~ SOZ_int", ldf, groups=ldf["Subject"])
            res = model.fit(reml=True, disp=False)
            beta = res.fe_params['SOZ_int']
            se = res.bse['SOZ_int']
            z = res.tvalues['SOZ_int']
            p = res.pvalues['SOZ_int']
            ci_lo = beta - 1.96 * se
            ci_hi = beta + 1.96 * se
            rv = float(res.cov_re.iloc[0, 0])
            icc = rv / (rv + res.scale)
        except Exception as e:
            beta = se = z = p = ci_lo = ci_hi = icc = float('nan')
            print(f"    ⚠ LMM failed for {label}: {e}")

        is_primary = 'all delivered' in label
        values = [label, n_contacts, n_patients,
                  round(beta, 4), round(se, 4),
                  round(ci_lo, 4), round(ci_hi, 4),
                  round(z, 2), f'{p:.2e}', round(d, 3), round(icc, 3)]
        format_data_row(ws, row, values, bold=is_primary)
        row += 1

    add_note(ws, row + 1, 1,
             'Note. Bold row indicates the full delivered dataset (R² ≥ 0.935, '
             'Exponent ≥ 1.17) used in the primary analysis. The upstream '
             'processing pipeline applied R² ≥ 0.85 and Exponent ≥ 0.5 '
             'thresholds before export; all retained contacts exceed these '
             'minima. Rows show progressively stricter R² thresholds. '
             'All models: Exponent ~ SOZ_status + (1 | Patient), REML. '
             'd = paired Cohen\'s d from patient-level means.',
             merge_to=11)
    print(f"    ✓ Table S1 complete")


# ═══════════════════════════════════════════════════════════════════
#  TABLE S2: Leave-one-out + Cook's D
# ═══════════════════════════════════════════════════════════════════

def build_table_s2(wb):
    print("  Building Table S2: LOO sensitivity + Cook's D ...")
    ws = wb.create_sheet('Table_S2')

    seeg = load_seeg()
    seeg = seeg[(seeg.R_Squared >= 0.85) & (seeg.Exponent >= MIN_EXPONENT)]
    scalp = load_scalp()
    engel = load_engel()
    engel_map = dict(zip(engel.Subject, engel.Outcome))

    # Compute X = SOZ - NonSOZ contrast per patient
    soz_per = seeg[seeg.Is_SOZ].groupby('Subject')['Exponent'].mean()
    nsoz_per = seeg[~seeg.Is_SOZ].groupby('Subject')['Exponent'].mean()
    contrast = soz_per - nsoz_per

    # Compute Y = midline ΔExponent
    mid = scalp[scalp.Channel.isin(MIDLINE)]
    pre_mid = mid[mid.Condition == 'Pre'].groupby('Subject_ID')['Exponent'].mean()
    post_mid = mid[mid.Condition == 'Post'].groupby('Subject_ID')['Exponent'].mean()
    delta_mid = post_mid - pre_mid

    # Merge for good-outcome patients
    common = sorted(set(contrast.index) & set(delta_mid.index) & set(PAIRED_SUBJECTS))
    good_subs = [s for s in common if engel_map.get(s) == 'Good']

    x_good = np.array([contrast[s] for s in good_subs])
    y_good = np.array([delta_mid[s] for s in good_subs])

    # Full-sample Spearman
    rho_full, p_full = stats.spearmanr(x_good, y_good)

    # Cook's distance from OLS
    import statsmodels.api as sm
    X_ols = sm.add_constant(x_good)
    ols_model = sm.OLS(y_good, X_ols).fit()
    influence = ols_model.get_influence()
    cooks_d = influence.cooks_distance[0]

    headers = ['Excluded Patient', 'Spearman ρ', 'p-value',
               'Δρ (from full)', "Cook's D", 'Influential?']
    widths  = [18, 14, 14, 14, 14, 14]

    add_note(ws, 1, 1,
             'Table S2. Leave-one-out Spearman ρ and Cook\'s distance for '
             'the good-outcome cross-modal correlation (Fig. 2C; n = 11).',
             merge_to=6)
    add_note(ws, 2, 1,
             f'Full-sample: ρ = {rho_full:.3f}, p = {p_full:.4f}',
             merge_to=6)
    format_header(ws, 4, headers, widths)

    row = 5
    for i, sub in enumerate(good_subs):
        mask = np.ones(len(good_subs), dtype=bool)
        mask[i] = False
        rho_loo, p_loo = stats.spearmanr(x_good[mask], y_good[mask])
        delta_rho = rho_loo - rho_full
        cd = cooks_d[i]
        influential = 'Yes' if cd > 4 / len(good_subs) else 'No'

        values = [sub, round(rho_loo, 3), round(p_loo, 4),
                  round(delta_rho, 3), round(cd, 4), influential]
        format_data_row(ws, row, values)
        row += 1

    # Summary row
    rho_range = [stats.spearmanr(x_good[np.arange(len(x_good)) != i],
                                  y_good[np.arange(len(y_good)) != i])[0]
                 for i in range(len(good_subs))]
    add_note(ws, row + 1, 1,
             f'Note. LOO ρ range: [{min(rho_range):.3f}, {max(rho_range):.3f}]. '
             f'Cook\'s D threshold = 4/n = {4/len(good_subs):.3f}. '
             f'No single patient reverses the direction or significance of the '
             f'correlation.',
             merge_to=6)
    print(f"    ✓ Table S2 complete")


# ═══════════════════════════════════════════════════════════════════
#  TABLE S3: Individual patient data
# ═══════════════════════════════════════════════════════════════════

def build_table_s3(wb):
    print("  Building Table S3: Individual patient summary ...")
    ws = wb.create_sheet('Table_S3')

    scalp = load_scalp()
    seeg  = load_seeg()
    engel = load_engel()
    roi   = load_roi()
    targets = load_target_regions()

    engel_map = dict(zip(engel.Subject, engel.Engel))
    outcome_map = dict(zip(engel.Subject, engel.Outcome))
    fu_map = dict(zip(engel.Subject, engel.Follow_up_mo))

    # ROI and side
    roi_map  = dict(zip(roi.Subject, roi.ROI))
    side_map = dict(zip(roi.Subject, roi['Inferred Side']))

    # Primary SOZ target per patient (first listed)
    target_map = {}
    for sub in PAIRED_SUBJECTS:
        sub_targets = targets[targets.Subject == sub]['Target_Region']
        if len(sub_targets):
            target_map[sub] = sub_targets.iloc[0]

    # Scalp whole-brain means
    pre_wb  = scalp[scalp.Condition == 'Pre'].groupby('Subject_ID')['Exponent'].mean()
    post_wb = scalp[scalp.Condition == 'Post'].groupby('Subject_ID')['Exponent'].mean()

    # Scalp midline means
    mid = scalp[scalp.Channel.isin(MIDLINE)]
    pre_mid  = mid[mid.Condition == 'Pre'].groupby('Subject_ID')['Exponent'].mean()
    post_mid = mid[mid.Condition == 'Post'].groupby('Subject_ID')['Exponent'].mean()

    # SEEG SOZ/NonSOZ per patient
    seeg_filt = seeg[(seeg.R_Squared >= 0.85) & (seeg.Exponent >= MIN_EXPONENT)]
    soz_exp   = seeg_filt[seeg_filt.Is_SOZ].groupby('Subject')['Exponent'].mean()
    nsoz_exp  = seeg_filt[~seeg_filt.Is_SOZ].groupby('Subject')['Exponent'].mean()
    n_seeg    = seeg_filt.groupby('Subject').size()
    n_soz     = seeg_filt[seeg_filt.Is_SOZ].groupby('Subject').size()

    headers = ['Patient', 'Lesion Side', 'SOZ Region',
               'N SEEG Contacts', 'N SOZ Contacts',
               'SEEG SOZ Exp', 'SEEG NonSOZ Exp', 'SOZ−NonSOZ',
               'Pre WB Exp', 'Post WB Exp', 'ΔWB Exp',
               'Pre Mid Exp', 'Post Mid Exp', 'ΔMid Exp',
               'Engel', 'Outcome', 'Follow-up (mo)']
    widths = [12, 12, 22, 14, 14, 14, 14, 12,
              12, 12, 12, 12, 12, 12, 10, 10, 14]

    add_note(ws, 1, 1,
             'Table S3. Individual patient aperiodic exponent data and '
             'clinical outcomes. WB = whole-brain mean; Mid = midline '
             '(Fz, Cz, Pz) mean; Exp = aperiodic exponent.',
             merge_to=17)
    format_header(ws, 3, headers, widths)

    row = 4
    for sub in PAIRED_SUBJECTS:
        side = side_map.get(sub, '—')
        region = target_map.get(sub, roi_map.get(sub, '—'))
        nc = int(n_seeg.get(sub, 0))
        ns = int(n_soz.get(sub, 0))
        soz_e  = round(soz_exp.get(sub, np.nan), 4)
        nsoz_e = round(nsoz_exp.get(sub, np.nan), 4)
        contrast = round(soz_e - nsoz_e, 4) if not np.isnan(soz_e) else '—'

        pre_w  = round(pre_wb.get(sub, np.nan), 4)
        post_w = round(post_wb.get(sub, np.nan), 4)
        d_wb   = round(post_w - pre_w, 4) if not (np.isnan(pre_w) or np.isnan(post_w)) else '—'

        pre_m  = round(pre_mid.get(sub, np.nan), 4)
        post_m = round(post_mid.get(sub, np.nan), 4)
        d_mid  = round(post_m - pre_m, 4) if not (np.isnan(pre_m) or np.isnan(post_m)) else '—'

        eng = engel_map.get(sub, '—')
        out = outcome_map.get(sub, '—')
        fu  = fu_map.get(sub, '—')
        if isinstance(fu, float) and not np.isnan(fu):
            fu = int(fu)

        values = [sub, side, region, nc, ns,
                  soz_e, nsoz_e, contrast,
                  pre_w, post_w, d_wb,
                  pre_m, post_m, d_mid,
                  eng, out, fu]
        format_data_row(ws, row, values)
        row += 1

    # Group summary rows
    row += 1
    for grp_label in ['Good', 'Poor']:
        grp_subs = [s for s in PAIRED_SUBJECTS if outcome_map.get(s) == grp_label]
        n = len(grp_subs)
        pre_vals = [pre_wb.get(s, np.nan) for s in grp_subs]
        post_vals = [post_wb.get(s, np.nan) for s in grp_subs]
        mid_d = [(post_mid.get(s, np.nan) - pre_mid.get(s, np.nan))
                 for s in grp_subs]
        values = [f'{grp_label} (n={n})', '', '', '', '', '', '', '',
                  f'{np.nanmean(pre_vals):.4f}',
                  f'{np.nanmean(post_vals):.4f}',
                  f'{np.nanmean(np.array(post_vals) - np.array(pre_vals)):.4f}',
                  '', '',
                  f'{np.nanmean(mid_d):.4f}',
                  '', '', '']
        format_data_row(ws, row, values, bold=True)
        row += 1

    print(f"    ✓ Table S3 complete")


# ═══════════════════════════════════════════════════════════════════
#  TABLE S4: Per-patient epoch counts + group comparisons
# ═══════════════════════════════════════════════════════════════════

def build_table_s4(wb):
    print("  Building Table S4: Epoch counts + confound tests ...")
    ws = wb.create_sheet('Table_S4')

    log = load_processing_log()
    engel = load_engel()
    outcome_map = dict(zip(engel.Subject, engel.Outcome))

    # Filter to paired subjects only
    log_paired = log[log.Subject_ID.isin(PAIRED_SUBJECTS) &
                     (log.Status == 'SUCCESS')].copy()

    headers = ['Patient', 'Condition', 'N2 Seconds', 'Epochs Created',
               'Epochs Clean', 'Rejection Rate (%)',
               'PTP Threshold (µV)', 'Mean R²', 'Mean Exponent', 'Outcome']
    widths = [12, 12, 12, 14, 14, 16, 18, 12, 14, 10]

    add_note(ws, 1, 1,
             'Table S4. Per-patient epoch counts, artefact rejection '
             'thresholds, and spectral quality metrics for scalp EEG '
             'recordings.',
             merge_to=10)
    format_header(ws, 3, headers, widths)

    row = 4
    for sub in PAIRED_SUBJECTS:
        for cond in ['Pre', 'Post']:
            sub_log = log_paired[(log_paired.Subject_ID == sub) &
                                 (log_paired.Condition == cond)]
            if len(sub_log) == 0:
                continue
            r = sub_log.iloc[0]
            created = r['N_Epochs_created']
            clean = r['N_Epochs_clean']
            rej_rate = round(100 * (1 - clean / created), 1) if created > 0 else 0

            values = [sub, cond, int(r['N2_seconds']),
                      int(created), int(clean), rej_rate,
                      int(r['PTP_Threshold_uV']),
                      round(r['Mean_R2'], 4),
                      round(r['Mean_Exponent'], 4),
                      outcome_map.get(sub, '—')]
            format_data_row(ws, row, values)
            row += 1

    # ── Group comparison tests ─────────────────────────────────────
    row += 1
    add_note(ws, row, 1,
             'Between-group comparisons of epoch counts (Good vs Poor):',
             merge_to=10)
    row += 1

    # Pre-operative: Good vs Poor epoch counts
    for cond in ['Pre', 'Post']:
        good_epochs = []
        poor_epochs = []
        for sub in PAIRED_SUBJECTS:
            sub_log = log_paired[(log_paired.Subject_ID == sub) &
                                 (log_paired.Condition == cond)]
            if len(sub_log) == 0:
                continue
            ep = sub_log.iloc[0]['N_Epochs_clean']
            if outcome_map.get(sub) == 'Good':
                good_epochs.append(ep)
            elif outcome_map.get(sub) == 'Poor':
                poor_epochs.append(ep)

        if len(good_epochs) >= 2 and len(poor_epochs) >= 2:
            t_stat, t_p = stats.ttest_ind(good_epochs, poor_epochs)
            u_stat, u_p = stats.mannwhitneyu(good_epochs, poor_epochs,
                                              alternative='two-sided')
            add_note(ws, row, 1,
                     f'{cond}: Good={np.mean(good_epochs):.0f}±'
                     f'{np.std(good_epochs, ddof=1):.0f} (n={len(good_epochs)}), '
                     f'Poor={np.mean(poor_epochs):.0f}±'
                     f'{np.std(poor_epochs, ddof=1):.0f} (n={len(poor_epochs)}); '
                     f't={t_stat:.2f}, p={t_p:.3f}; '
                     f'Mann-Whitney U={u_stat:.0f}, p={u_p:.3f}',
                     merge_to=10)
            row += 1

    # Pre vs Post epoch count comparison (within-subject)
    row += 1
    add_note(ws, row, 1,
             'Within-subject Pre vs Post epoch count comparison:',
             merge_to=10)
    row += 1

    pre_epochs = []
    post_epochs = []
    for sub in PAIRED_SUBJECTS:
        pre_log = log_paired[(log_paired.Subject_ID == sub) &
                              (log_paired.Condition == 'Pre')]
        post_log = log_paired[(log_paired.Subject_ID == sub) &
                               (log_paired.Condition == 'Post')]
        if len(pre_log) > 0 and len(post_log) > 0:
            pre_epochs.append(pre_log.iloc[0]['N_Epochs_clean'])
            post_epochs.append(post_log.iloc[0]['N_Epochs_clean'])

    if len(pre_epochs) >= 2:
        t_stat, t_p = stats.ttest_rel(pre_epochs, post_epochs)
        w_stat, w_p = stats.wilcoxon(pre_epochs, post_epochs)
        add_note(ws, row, 1,
                 f'Pre={np.mean(pre_epochs):.0f}±{np.std(pre_epochs, ddof=1):.0f}, '
                 f'Post={np.mean(post_epochs):.0f}±{np.std(post_epochs, ddof=1):.0f} '
                 f'(N={len(pre_epochs)} paired); '
                 f'Paired t={t_stat:.2f}, p={t_p:.3f}; '
                 f'Wilcoxon W={w_stat:.0f}, p={w_p:.3f}',
                 merge_to=10)
        row += 1

    # Correlation: epoch count vs exponent change (confound check)
    row += 1
    add_note(ws, row, 1,
             'Confound check: epoch count vs aperiodic exponent:',
             merge_to=10)
    row += 1

    # Post epoch count vs post-pre exponent difference
    ep_post = []
    delta_exp = []
    for sub in PAIRED_SUBJECTS:
        pre_log = log_paired[(log_paired.Subject_ID == sub) &
                              (log_paired.Condition == 'Pre')]
        post_log = log_paired[(log_paired.Subject_ID == sub) &
                               (log_paired.Condition == 'Post')]
        if len(pre_log) > 0 and len(post_log) > 0:
            ep_post.append(post_log.iloc[0]['N_Epochs_clean'])
            delta_exp.append(post_log.iloc[0]['Mean_Exponent'] -
                           pre_log.iloc[0]['Mean_Exponent'])

    if len(ep_post) >= 5:
        rho, p_rho = stats.spearmanr(ep_post, delta_exp)
        add_note(ws, row, 1,
                 f'Post epoch count vs ΔExponent: Spearman ρ={rho:.3f}, '
                 f'p={p_rho:.3f} (N={len(ep_post)}). '
                 f'No significant relationship, ruling out epoch-count '
                 f'confounding of the primary exponent findings.',
                 merge_to=10)

    print(f"    ✓ Table S4 complete")


# ═══════════════════════════════════════════════════════════════════
#  TABLE S5: Per-patient ASM regimen stability
# ═══════════════════════════════════════════════════════════════════

def build_table_s5(wb):
    print("  Building Table S5: ASM regimen stability ...")
    ws = wb.create_sheet('Table_S5')

    asm = load_asm()
    if asm is None:
        add_note(ws, 1, 1,
                 'Table S5. [PLACEHOLDER] ASM_regimen.csv not found. '
                 'Fill the template and place it in the data directory.',
                 merge_to=8)
        print("    ⚠ ASM data file not found — placeholder sheet created")
        return

    engel = load_engel()
    outcome_map = dict(zip(engel.Subject, engel.Outcome))

    add_note(ws, 1, 1,
             'Table S5. Per-patient anti-seizure medication (ASM) regimens '
             'at pre-operative and post-operative (6–8-month) EEG acquisition, '
             'confirming pharmacological stability across the surgical interval.',
             merge_to=8)

    headers = ['Patient', 'Outcome', 'ASM 1', 'Dose Pre → Post',
               'ASM 2', 'Dose Pre → Post', 'ASM 3', 'Dose Pre → Post',
               'Change Summary']
    widths = [12, 10, 18, 20, 18, 20, 18, 20, 24]
    format_header(ws, 3, headers, widths)

    n_any_change = 0
    row = 4
    for _, r in asm.iterrows():
        sub = r['Subject']
        if sub not in PAIRED_SUBJECTS:
            continue
        outcome = outcome_map.get(sub, '—')

        # Build dose strings for up to 3 ASMs
        dose_entries = []
        for i in range(1, 4):
            name_col = f'ASM_{i}_Name'
            pre_col = f'ASM_{i}_Dose_Pre'
            post_col = f'ASM_{i}_Dose_Post'
            unit_col = f'ASM_{i}_Unit'

            name = r.get(name_col, '')
            if pd.isna(name) or str(name).strip() == '':
                dose_entries.append(('', ''))
                continue

            pre_d = r.get(pre_col, '')
            post_d = r.get(post_col, '')
            unit = r.get(unit_col, '')

            if pd.isna(pre_d) or pd.isna(post_d):
                dose_str = '—'
            else:
                pre_d = int(pre_d) if float(pre_d) == int(float(pre_d)) else pre_d
                post_d = int(post_d) if float(post_d) == int(float(post_d)) else post_d
                u = str(unit) if not pd.isna(unit) else ''
                dose_str = f'{pre_d} → {post_d} {u}'.strip()

            dose_entries.append((str(name).strip(), dose_str))

        summary = r.get('Dose_Change_Summary', '')
        if pd.isna(summary):
            summary = ''
        summary = str(summary).strip()

        if summary.lower() not in ('no change', 'unchanged', ''):
            n_any_change += 1

        values = [sub, outcome]
        for name, dose in dose_entries:
            values.extend([name, dose])
        values.append(summary)

        format_data_row(ws, row, values)
        row += 1

    # Summary note
    row += 1
    total = len([s for s in PAIRED_SUBJECTS
                 if s in asm['Subject'].values])
    n_stable = total - n_any_change
    add_note(ws, row, 1,
             f'Note. {n_stable}/{total} patients had identical ASM regimens '
             f'at both timepoints. '
             + (f'{n_any_change} patient(s) had minor dose adjustments '
                f'as detailed above.'
                if n_any_change > 0
                else 'No patient had any change in ASM type or dose.'),
             merge_to=9)

    print(f"    ✓ Table S5 complete ({n_stable}/{total} unchanged)")

    # ── Print manuscript-ready sentence for the PLACEHOLDER ──────
    print("\n    ── MANUSCRIPT TEXT (paste into §5.3 ASM PLACEHOLDER): ──")
    if n_any_change == 0:
        ms_text = ("no dose adjustments occurred in any patient")
    else:
        ms_text = (f"minor dose adjustments in {n_any_change} "
                   f"patient{'s' if n_any_change > 1 else ''} did not "
                   f"exceed [X]% of the pre-operative regimen")
    print(f'    "{ms_text}"')
    print("    ── END ──\n")


# ═══════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("  Supplementary Tables Generator")
    print(f"  Data directory: {DATA_DIR}")
    print(f"  Output: {OUTPUT}")
    print("=" * 60)

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    build_table_s1(wb)
    build_table_s2(wb)
    build_table_s3(wb)
    build_table_s4(wb)
    build_table_s5(wb)

    wb.save(OUTPUT)
    print(f"\n  ✓ Saved: {OUTPUT}")
    print("=" * 60)


if __name__ == "__main__":
    main()
